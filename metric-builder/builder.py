"""
Core engine: config → validate → inject data → generate formulas → validate output.
"""
import csv
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

from .config import load_config
from .formulas import generate_rollup, generate_ranked, generate_headers
from .schema import validate_schema
from .validation import validate_completeness


class MetricBuilder:

    def __init__(self, config_path, data_path, template_path):
        self.config = load_config(config_path)
        self.data_path = data_path
        self.template_path = template_path

    def build(self, output_path, validate_only=False):
        data_rows = self._load_csv()
        errors = validate_schema(data_rows, self.config)
        if errors:
            return {"status": "error", "stage": "schema", "errors": errors}

        if validate_only:
            return {"status": "ok", "rows": len(data_rows)}

        wb = openpyxl.load_workbook(str(self.template_path))
        self._inject_data(wb, data_rows)
        generate_headers(wb, self.config, data_rows)

        for section in self.config.sections:
            if section.type == "rollup":
                generate_rollup(wb, section, self.config)
            elif section.type == "ranked":
                generate_ranked(wb, section, self.config)

        warnings = validate_completeness(wb, self.config, data_rows)
        wb.save(str(output_path))
        return {"status": "ok", "output": str(output_path), "warnings": warnings}

    def _load_csv(self):
        with open(self.data_path, newline='') as f:
            return list(csv.DictReader(f))

    def _inject_data(self, wb, data_rows):
        tab = self.config.data.tab_name
        if tab not in wb.sheetnames:
            wb.create_sheet(tab)
        ws = wb[tab]
        headers = list(data_rows[0].keys())
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, row in enumerate(data_rows, 2):
            for col, h in enumerate(headers, 1):
                val = row[h]
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    pass
                ws.cell(row=row_idx, column=col, value=val)
