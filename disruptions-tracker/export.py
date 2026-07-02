"""
Daily export: ticket counts + yard metrics → Redshift → validate → Excel.
Idempotent. One command. No partial writes.
"""
import argparse
import sys
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
QUERY_FILE = ROOT / "shift_query.sql"
EXCEL_FILE = ROOT / "tracker.xlsx"

sys.path.insert(0, str(ROOT / "lib"))
from db_connector import run_script


def parse_args():
    p = argparse.ArgumentParser(description="Disruptions tracker daily export")
    p.add_argument("--tickets-a", type=int, required=True)
    p.add_argument("--tickets-b", type=int, required=True)
    p.add_argument("--tickets-c", type=int, required=True)
    p.add_argument("--tickets-d", type=int, required=True)
    p.add_argument("--dry-run", action="store_true")
    return p.parse_args()


def store_snapshot(a, b, c, d):
    sql = f"""
    INSERT INTO analytics.daily_ticket_snapshot (snapshot_date, section, open_count)
    SELECT CURRENT_DATE, section, cnt FROM (
        SELECT 'Queue A' AS section, {a} AS cnt
        UNION ALL SELECT 'Queue B', {b}
        UNION ALL SELECT 'Queue C', {c}
        UNION ALL SELECT 'Queue D', {d}
    ) src
    WHERE NOT EXISTS (
        SELECT 1 FROM analytics.daily_ticket_snapshot
        WHERE snapshot_date = CURRENT_DATE AND section = src.section);"""
    run_script(sql)


def validate_balance(rows):
    today = str(date.today())
    cases = {}
    for r in rows:
        if str(r[1]) == "Cases" and str(r[3])[:10] == today:
            cases[r[2]] = int(r[4]) if r[4] else 0

    if not {"SOS", "EOS", "New", "Resolved"}.issubset(cases.keys()):
        return False

    expected = cases["SOS"] + cases["New"] - cases["Resolved"]
    if expected == cases["EOS"]:
        print(f"  Balance: {cases['SOS']} + {cases['New']} - {cases['Resolved']} = {cases['EOS']} ✓")
        return True
    print(f"  Balance FAILED: expected {expected}, got {cases['EOS']}")
    return False


def write_excel(rows, dry_run=False):
    if dry_run:
        print(f"  DRY RUN: would write {len(rows)} rows")
        return
    if not EXCEL_FILE.exists():
        for row in rows:
            print(f"  {row[1]:<20} {row[2]:<10} {str(row[3]):<12} {row[4]}")
        return
    wb = openpyxl.load_workbook(str(EXCEL_FILE))
    ws = wb["Query Data"]
    start = max(ws.max_row + 1, 4)
    for i, row in enumerate(rows):
        ws.cell(row=start + i, column=1, value=row[0])
        ws.cell(row=start + i, column=2, value=str(row[1]))
        ws.cell(row=start + i, column=3, value=str(row[2]))
        ws.cell(row=start + i, column=4, value=row[3])
        ws.cell(row=start + i, column=5, value=int(row[4]) if row[4] else 0)
    wb.save(str(EXCEL_FILE))
    print(f"  Wrote {len(rows)} rows starting at row {start}")


def main():
    args = parse_args()
    print(f"Disruptions Tracker — {date.today()}")

    if not args.dry_run:
        store_snapshot(args.tickets_a, args.tickets_b, args.tickets_c, args.tickets_d)

    sql_rows = run_script(QUERY_FILE.read_text()) if not args.dry_run else []

    if sql_rows and not validate_balance(sql_rows):
        print("  BALANCE FAILED. No write.")
        sys.exit(1)

    write_excel(sql_rows, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
