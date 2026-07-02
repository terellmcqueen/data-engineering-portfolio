"""Cell-level diff of two Excel workbooks. Handles merged cells, numeric tolerance."""
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

DEFAULT_TOLERANCE = 1e-9


@dataclass
class CellChange:
    sheet: str
    cell: str
    change_type: str  # "modified", "added", "removed"
    old_value: Any = None
    new_value: Any = None


@dataclass
class DiffSummary:
    total_changes: int = 0
    sheets_affected: int = 0
    added: int = 0
    modified: int = 0
    removed: int = 0


def diff_workbooks(path_old, path_new, tolerance=DEFAULT_TOLERANCE):
    wb_old = load_workbook(str(path_old), data_only=True)
    wb_new = load_workbook(str(path_new), data_only=True)
    changes = []

    all_sheets = set(wb_old.sheetnames) | set(wb_new.sheetnames)
    for sheet in all_sheets:
        if sheet not in wb_old.sheetnames:
            changes.append(CellChange(sheet, "SHEET", "added"))
            continue
        if sheet not in wb_new.sheetnames:
            changes.append(CellChange(sheet, "SHEET", "removed"))
            continue

        ws_old, ws_new = wb_old[sheet], wb_new[sheet]
        max_row = max(ws_old.max_row or 0, ws_new.max_row or 0)
        max_col = max(ws_old.max_column or 0, ws_new.max_column or 0)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                old_val = ws_old.cell(row=row, column=col).value
                new_val = ws_new.cell(row=row, column=col).value

                if _values_differ(old_val, new_val, tolerance):
                    cell_ref = ws_old.cell(row=row, column=col).coordinate
                    if old_val is None:
                        changes.append(CellChange(sheet, cell_ref, "added", None, new_val))
                    elif new_val is None:
                        changes.append(CellChange(sheet, cell_ref, "removed", old_val, None))
                    else:
                        changes.append(CellChange(sheet, cell_ref, "modified", old_val, new_val))

    summary = DiffSummary(
        total_changes=len(changes),
        sheets_affected=len(set(c.sheet for c in changes)),
        added=sum(1 for c in changes if c.change_type == "added"),
        modified=sum(1 for c in changes if c.change_type == "modified"),
        removed=sum(1 for c in changes if c.change_type == "removed"),
    )
    return changes, summary


def _values_differ(a, b, tolerance):
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) > tolerance
    return str(a) != str(b)
